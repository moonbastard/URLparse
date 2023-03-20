from flask import Flask, render_template, request
import re
import requests
from bs4 import BeautifulSoup
import openpyxl

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/search', methods=['POST'])
def search():
    # Get the user input from the form
    search_phrase = request.form['search_phrase']
    url_prefix = request.form['url_prefix']
    url_list = request.files['url_list']

    # Read the URLs from the uploaded file
    wb = openpyxl.load_workbook(url_list)
    ws = wb.active
    urls = [url_prefix + cell.value for row in ws.iter_rows() for cell in row if cell.value]

    # Define the regular expression to match the search phrase
    search_regex = re.compile(search_phrase, re.IGNORECASE)

    # Create a new Excel document
    wb = openpyxl.Workbook()
    ws = wb.active

    # Define the column headers
    ws.cell(1, 1).value = "URL"
    ws.cell(1, 2).value = "Line"

    # Keep track of the current row to write to
    row_num = 2

    # Iterate over each URL
    for url in urls:
        # Make a request to the URL
        res = requests.get(url)

        # Parse the HTML content
        soup = BeautifulSoup(res.text, 'html.parser')

        # Find the <body> tag
        body = soup.find('body')

        # Iterate over each line in the <body> tag
        for line in body.text.splitlines():
            # Check if the line contains the search phrase
            match = search_regex.search(line)
            if match:
                # Write the URL to the first column
                ws.cell(row_num, 1).value = url

                # Write the line containing the search phrase to the second column
                ws.cell(row_num, 2).value = line

                # Increment the row number
                row_num += 1

    # Save the Excel document
    wb.save('results.xlsx')

    # Return the results page
    return render_template('results.html')
